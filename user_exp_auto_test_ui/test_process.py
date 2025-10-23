from dataclasses import dataclass, asdict
from bt_control import BluetoothControl as b_control
from pynput.keyboard import Key, Controller
from teams_meeting_control import MeetingControl
from teams_call.auto_vpt.vpt_contoller import VPTControl
from wrt_controller import WRTController
from audio_detect_control import AudioDetectController
from video_control import VideoControl
from utils import Utils
from enum import Enum
import serial.tools.list_ports
from serial.serialutil import SerialException
import time
import serial
import os
import pygame
import pyautogui
from pynput import mouse, keyboard
from datetime import datetime
import yaml
from utils import log as logger
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from youtube_control import YoutubeControl
import pygetwindow as gw
from database_manager import Database_data


class Power_States(Enum):
    idle = 0
    go_to_s3 = 1
    go_to_s4 = 2


class Headset(Enum):
    idle = 0
    turn_on_off = 1


class ENV(Enum):
    Teams = 0
    Local_audio = 1
    Local_video = 2
    Teams_Local = 3
    Youtube = 4


class Test_case(Enum):
    Idle = "Idle"
    MS = "Modern_Standby"
    S4 = "Hibernation"
    Mouse_function = "Mouse_Function_Check"
    keyboard_function = "Keyboard_Function_Check"
    keyboard_latency = "Keyboard_Latency_Test"
    Keyboard_random = "Keyboard_Random_Click"
    Mouse_random = "Mouse_Random_Click"
    Headset_init = "Headset_Initialization"
    Headset_input = "Headset_Mic_Check"
    Headset_output = "Headset_Output_Check"
    Headset_del = "Headset_Restore"
    Mouse_latency = "Mouse_Latency_Test"
    Environment_init = "Environment_Initialization"
    Environment_restore = "Environment_Restore"


# Arduino cmd setting
CMD_servo = str.encode("0")
CMD_voice_detect = str.encode("1")
CMD_buzzer = str.encode("2")
CMD_mouse_clicking = str.encode("3")
CMD_mouse_delay_clicking = str.encode("4")
CMD_mouse_random_clicking = str.encode("5")
CMD_keyboard_clicking = str.encode("6")
CMD_mouse_latency = str.encode("7")
CMD_keyboard_latency = str.encode("8")
CMD_keyboard_random_clicking = str.encode("9")
CMD_test = str.encode("f")
g_COM_PORT = ""


@dataclass
class Basic_Config:
    headset: str = "Zone"
    target_port_desc: str = "USB-SERIAL CH340"
    teams_url: str = (
        "https://teams.microsoft.com/l/meetup-join/19%3ameeting_MWQ5MTEwZmUtNWZkMy00YzZkLTgwOTMtNWVjMTc3NjMxZWMz%40thread.v2/0?context=%7b%22Tid%22%3a%228e44b933-0b1e-4e67-be0c-c7e0761eb4db%22%2c%22Oid%22%3a%2232692ba0-ede0-450b-8142-3f487cabff7b%22%7d"
    )
    timeout_s: int = 5
    sleep_time_s: int = 30
    wake_up_time_s: int = 60
    ENV_source: int = ENV.Local_audio.value
    headset_setting: int = Headset.idle.value
    test_retry_times: int = 3
    continue_fail_limit: int = 5
    output_source_play_time_s: int = 20
    task_schedule: str = "MS,Idle,headset_output"
    test_times: int = 100
    com: str = ""
    youtube_link: str = "https://www.youtube.com/watch?v=w9k7eWD0ik8"
    mouse_latency_threshold: int = 80
    keyboard_latency_threshold: int = 100
    report_path:str = ""


def load_basic_config(file_path: str) -> Basic_Config:
    with open(file_path) as f:
        data = yaml.safe_load(f)
    return Basic_Config(**data)


def create_report_folder(base_path="report") -> str:
    # Folder name with current date
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder_path = os.path.join(base_path, f"Report_{date_str}")

    # Create folder if it doesn't exist
    os.makedirs(folder_path, exist_ok=True)
    print(f"Report folder created at: {folder_path}")

    return folder_path


def save_report(
    config: Basic_Config,
    data: Database_data,
    total_cycles: int,
    fail_times: int,
    error_message="",
    
):
    wb = Workbook()
    ws = wb.active
    ws.title = "BT Test Report"

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Bluetooth Device Test Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Summary Section
    ws.append([])
    ws.append(["Test Summary"])
    ws["A3"].font = Font(size=14, bold=True)
    ws.append(["Total Test Cycles", total_cycles])
    ws.append(["Pass Times", (total_cycles - fail_times)])
    ws.append(["Fail Times", fail_times])
    ws.append(["Error Message", error_message])

    # Config Section
    ws.append([])
    ws.append(["Configuration Settings"])
    ws["A9"].font = Font(size=14, bold=True)
    row_start = 10
    for key, value in asdict(config).items():
        if isinstance(value, Enum):
            value = value.name
        ws.append([key, str(value)])
        row_start += 1

    # Database Section
    ws.append([])
    ws.append(["Database data"])
    ws["A{}".format(row_start + 1)].font = Font(size=14, bold=True)
    for key, value in asdict(data).items():
        ws.append([key, str(value)])

    # Make it look cleaner
    for col in ["A", "B"]:
        ws.column_dimensions[col].width = 25

    current_time = datetime.now()
    timestamp_str = current_time.strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(config.report_path,f"test_report_{timestamp_str}.xlsx")

    wb.save(filename)
    print(f"✅ Report saved as {filename}")


def go_to_sleep():
    """
    using keybaord to go to sleep mode
    """
    keyboard = Controller()
    keyboard.press(Key.cmd)
    time.sleep(0.3)
    keyboard.press("x")
    time.sleep(0.3)
    keyboard.release("x")
    time.sleep(0.3)
    keyboard.release(Key.cmd)
    time.sleep(0.3)
    keyboard.tap("u")
    time.sleep(0.3)
    keyboard.tap("s")


def arduino_board_check(target_port_desc: str) -> str:
    """
    arduino board checking
    """
    ports = serial.tools.list_ports.comports()
    for port, desc, hwid in ports:
        if target_port_desc in desc:
            target_port = port
            print(f"{port}: {desc}")
            return port
    if not target_port:
        return ""


def headset_init(
    headset_states: Headset,
    ser: serial.Serial,
    command: bytes,
    headset: str,
    timeout_s: int,
) -> bool:
    """
    headset states init and check
    idle: only check the headset states, don't turn on off the headset power
    turn_on_off: turn on the headset and then check the connect states
    """
    counter = 0

    match headset_states:
        case Headset.turn_on_off.value:
            while counter < timeout_s:
                if b_control.status_check(target=headset, type="AudioEndpoint"):
                    print("Headset connect!")
                    return True
                else:
                    print("headset doesn't connet, click the power button to turn on")
                    ser.write(command + b"\n")
                    time.sleep(12)
                    counter += 1
            return False
        case Headset.idle.value:
            return b_control.status_check(target=headset, type="AudioEndpoint")


def headset_del(
    headset_states: Headset,
    ser: serial.Serial,
    command: bytes,
    headset: str,
    timeout_s: int,
) -> bool:
    """
    idle: do nothing
    turn on off: turn off the headset and make sure the headset is disconnect
    """
    match headset_states:
        case Headset.turn_on_off.value:
            counter = 0
            while counter < timeout_s:
                if not b_control.status_check(target=headset, type="AudioEndpoint"):
                    print("Headset disconnect!")
                    return True
                else:
                    print("headset still, click the power button to turn off")
                    ser.write(command + b"\n")
                    time.sleep(10)
                    counter += 1
            return False
        case Headset.idle.value:
            return True


def voice_detect(ser: serial.Serial, command: bytes) -> bool:
    """
    detect the sound
    """
    ser.write(command + b"\n")
    time.sleep(3)
    res = ser.read()
    print(f"*******************res = {res}***********************")
    return True if b"1" in res else False


def buzzer_buzzing(ser: serial.Serial, command: bytes) -> bool:
    """
    let buzzer start buzzing for 5 sec
    """
    ser.write(command + b"\n")


def open_teams_call_and_join_meeting(t_control: MeetingControl) -> bool:
    """
    open teams call and join the specific meeting
    """
    t_control.open_teams()
    time.sleep(10)
    if t_control.join_meeting():
        time.sleep(5)
        # t_control.open_camera_and_mute()
        return True
    else:
        return False


def close_teams_call_and_vpt():
    """
    close the teams call and vpt robot
    """
    MeetingControl.end_meeting()
    time.sleep(5)
    MeetingControl.close_teams()
    VPTControl.vpt_bot_close()


def get_arduino_port(port: str, log_callback) -> str:
    """
    check and get the arduino board serial port
    """
    serial_port = ""
    try:
        serial_port = arduino_board_check(target_port_desc=port)
        log_callback(f"serial port detected: {serial_port}", False)
    except Exception as e:
        log_callback("please insert the arduino board!", False)

    return serial_port


def env_init(
    ENV_source: ENV, t_control: MeetingControl, v_control: VPTControl, log_callback
) -> bool:
    """
    doing some pre init before test the headset output function , like open teams or local music
    """
    match ENV_source:
        case ENV.Teams.value:
            # open the teams call and join the meeting
            res = open_teams_call_and_join_meeting(t_control=t_control)
            if not res:
                log_callback("Can not join the Teams call meeting!", False)
                close_teams_call_and_vpt()
                return False
            log_callback("Start the teams meeting...", False)
            # teams call robot join
            res = v_control.vpt_bot_join()
            if not res:
                log_callback("Can not execute VPT successfully", False)
                close_teams_call_and_vpt()
                return False
            log_callback("VPT teams robot join the meeting...", False)
            time.sleep(10)
        case ENV.Local_audio.value:
            log_callback("Start playing the local music...", False)
            videoControl = VideoControl(
                path=os.path.join("\\", "local_music", "test.mp3")
            )
            videoControl.play()
        case ENV.Local_video.value:
            log_callback("Start playing the local music...", False)
            videoControl = VideoControl(
                path=os.path.join("\\", "local_music", "test.mp4")
            )
            videoControl.play()
        case ENV.Teams_Local.value:
            # open the teams call and join the meeting
            res = open_teams_call_and_join_meeting(t_control=t_control)
            if not res:
                log_callback("Can not join the Teams call meeting!", False)
                close_teams_call_and_vpt()
                return False
            log_callback("Start the teams meeting...", False)
            # start playing local music after joining the teams call meeting
            log_callback("Start playing the local music...", False)
            videoControl = VideoControl(
                path=os.path.join("\\", "local_music", "test.mp4")
            )
            videoControl.play()
        case ENV.Youtube.value:
            log_callback("Start Youtube...", False)
            youtubeControl = YoutubeControl(
                link="https://www.youtube.com/watch?v=w9k7eWD0ik8"
            )
            youtubeControl.play()
    time.sleep(5)
    # make sure the main ui will be on the top of the screen
    window: gw.Win32Window
    windows = gw.getWindowsWithTitle("User Experience Auto Test")
    window = windows[0]
    window.minimize()
    window.restore()
    window.moveTo(0, 0)

    return True


def headset_output_test(
    b_config: Basic_Config, ser: serial.Serial, log_callback
) -> bool:
    log_callback("Start headset output function test...", False)
    test_time = 0
    res_output = False

    for _ in range(50):
        pyautogui.press("volumeup")
        time.sleep(0.1)

    while not res_output:
        # start play sound before detect
        time.sleep(b_config.output_source_play_time_s)
        # voice detect
        res_output = voice_detect(ser=ser, command=CMD_voice_detect)

        if not res_output:
            log_callback("headset output test fail, retry again!", False)
        if test_time > b_config.test_retry_times:
            log_callback("***Headset output function have some issue!***", False)
            log_callback("Dump WRT log...", False)
            WRTController.dump_wrt_log(log_path=b_config.report_path)
            return False
        test_time += 1
    log_callback("Headset output function test finish", False)
    return res_output


def headset_input_test(
    b_config: Basic_Config, ser: serial.Serial, log_callback
) -> bool:
    log_callback("Start headset input function test...", False)
    # headset input function test
    test_time = 0
    res_input = False
    while not res_input:
        ad_Controller = AudioDetectController(headset=b_config.headset, threshold=150)
        buzzer_buzzing(ser=ser, command=CMD_buzzer)
        res_input = ad_Controller.audio_detect()
        if test_time > b_config.test_retry_times:
            log_callback("***Headset input function have some issue!***", False)
            log_callback("Dump WRT log...", False)
            WRTController.dump_wrt_log(log_path=b_config.report_path)
            break
        test_time += 1
    log_callback("Headset input function test finish", False)
    time.sleep(5)
    return res_input


def env_restore(env_source: ENV, log_callback) -> bool:
    """
    doing some del before test the headset output function , like close teams or local music
    """
    match env_source:

        case ENV.Teams.value:
            # close the teams call and vpt robot
            close_teams_call_and_vpt()
            log_callback("End the teams meeting and VPT robot", False)

        case ENV.Local_audio.value:
            # close the media player
            VideoControl.stop_play()
            log_callback("End local music playing", False)

        case ENV.Local_video.value:
            # close the media player
            VideoControl.stop_play()
            log_callback("End local video playing", False)

        case ENV.Teams_Local.value:
            MeetingControl.close_teams()
            log_callback("End the teams meeting", False)
            VideoControl.stop_play()
            log_callback("End local music playing", False)

        case ENV.Youtube.value:
            YoutubeControl.Close()
            log_callback("End Youtube playing", False)

    return True


def mouse_keyboard_function_detect(
    ser: serial.Serial, command: bytes, timeout_s: int, log_callback
) -> bool:
    """_summary_
        check mouse function still working or not
    Returns:
        bool: _description_
    """
    counter = 0
    pygame.init()
    screen_width, screen_height = pyautogui.size()
    pygame.display.set_mode((screen_width, screen_height))

    # control mouse clicking
    ser.write(command + b"\n")
    # start cehcking mouse click
    while counter < timeout_s:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                log_callback("mouse function test have unexpected issue!", False)
                pygame.quit()
                return False
            elif event.type == pygame.MOUSEBUTTONDOWN:
                log_callback("BLE mouse function pass!", False)
                pygame.quit()
                return True
            elif event.type == pygame.KEYDOWN:
                log_callback("BLE keyboard function pass!", False)
                pygame.quit()
                return True

        counter += 1
        time.sleep(1)
    pygame.quit()
    return False


def mouse_keyboard_random_click(
    ser: serial.Serial, command: bytes, timeout_s: int, log_callback
) -> bool:
    if command == CMD_mouse_random_clicking:
        log_callback("mouse random click start!", False)
    else:
        log_callback("keyboard random click start!", False)

    # control mouse randomly clicking
    ser.write(command + b"\n")
    # waiting for the mouse click end
    ser.read()

    if command == CMD_keyboard_random_clicking:
        log_callback("mouse random click finish!", False)
    else:
        log_callback("keyboard random click finish!", False)
    return True


def mouse_function_detect_s3(
    ser: serial.Serial, command: bytes, sleep_time: int
) -> bool:
    """_summary_
        check mouse function still working or not (s3 states)
    Returns:
        bool: _description_
    """
    print("BLE mouse function test setting...")
    cmd = f"4,{sleep_time}\n".encode()
    ser.write(cmd)
    print(f"BLE mouse will click after {sleep_time} second! ")
    return True


def arduino_serial_port_reset(ser: serial.Serial, serial_port: str):
    """_summary_
    reset the arduino serial port after s3 or s4
    Args:
        ser (serial.Serial): _description_

    Returns:
        _type_: _description_
    """
    try:
        ser.close()
    except:
        pass
    time.sleep(5)
    return serial.Serial(serial_port, 115200)


def wait_for_port(port_name, timeout=30):
    """Wait until a specific COM port reappears"""
    start = time.time()
    while time.time() - start < timeout:
        ports = [p.device for p in serial.tools.list_ports.comports()]
        print(ports)
        if port_name in ports:
            return True
        time.sleep(1)
    return False


def dut_states_init(
    power_states: Power_States,
    wake_up_time_s: int,
    sleep_time_s: int,
    ser: serial.Serial,
) -> None:
    """_summary_
    setting dut states, s3 ,s4 or idle

    Args:
        states (States): _description_
    """
    match power_states:
        case Power_States.idle:
            time.sleep(wake_up_time_s)
        case Power_States.go_to_s3:
            mouse_function_detect_s3(
                ser=ser, command=CMD_mouse_delay_clicking, sleep_time=sleep_time_s
            )
            go_to_sleep()
        case Power_States.go_to_s4:
            # go to sleep mode first and waiting for mouse click
            cmd_pwrtest = f"pwrtest\pwrtest.exe /sleep /s:4 /c:1 /d:{sleep_time_s} /p:{wake_up_time_s}"
            Utils.run_sync_cmd(cmd=cmd_pwrtest)


def safe_write(ser: serial.Serial, data, baudrate=115200):
    global g_COM_PORT
    # safe serial wirte cmd to aviod issue
    try:
        ser.write(data)
    except SerialException as e:
        print(f"[WARN] Serial exception: {e}")
        try:
            ser.close()
        except:
            pass

        time.sleep(1)  # wait a bit for device to re-enumerate
        ser = serial.Serial(g_COM_PORT, baudrate, timeout=30)
        ser.write(data)
        ser.close()


g_latency = 0.0


def mouse_latency(
    ser: serial.Serial, threshold: int, timeout_s: int, log_callback=None
) -> bool:

    global g_latency

    def on_click(x, y, button, pressed):
        global g_latency
        if pressed:
            end = time.perf_counter()
            # minus the servo motor moving time
            g_latency = (end - start) - 0.173 - 0.5
            return False  # stop listener

    latency_list = []
    for _ in range(15):
        start = time.perf_counter()  # mark the time
        ser.write(CMD_mouse_latency + b"\n")  # example: send 'C' = click command
        ser.flush()
        with mouse.Listener(on_click=on_click) as listener:
            listener.join()

        if g_latency > 0:
            latency_list.append(g_latency * 1000)
            print(f"mouse clicking latency:{g_latency*1000:.3f} ms")
        else:
            print("retry...")
        time.sleep(3)

    if len(latency_list) > 4:
        for _ in range(2):
            min_val = min(latency_list)
            max_val = max(latency_list)
            latency_list.remove(min_val)
            latency_list.remove(max_val)
    else:
        log_callback("mouse latency Insuffcient sampling!", False)
        return False

    average_latency = sum(latency_list) / len(latency_list)
    log_callback(f"mouse average clicking latency:{average_latency:.3f} ms", False)

    return True if average_latency <= threshold else False


def keyboard_latency(
    ser: serial.Serial, threshold: int, timeout_s: int, log_callback=None
) -> bool:

    global g_latency

    def on_press(key):
        global g_latency
        end = time.perf_counter()
        # minus the servo motor moving time
        g_latency = (end - start) - 0.166 - 0.5
        return False  # stop listener

    latency_list = []
    for _ in range(15):
        start = time.perf_counter()  # mark the time
        ser.write(CMD_keyboard_latency + b"\n")  # example: send 'C' = click command
        ser.flush()
        with keyboard.Listener(on_press=on_press) as listener:
            listener.join()

        if g_latency > 0:
            latency_list.append(g_latency * 1000)
            print(f"keyboard clicking latency:{g_latency*1000:.3f} ms")
        else:
            print("retry...")
        time.sleep(3)

    if len(latency_list) > 4:
        for _ in range(2):
            min_val = min(latency_list)
            max_val = max(latency_list)
            latency_list.remove(min_val)
            latency_list.remove(max_val)
    else:
        log_callback("keyboard latency Insuffcient sampling!", False)
        return False

    average_latency = sum(latency_list) / len(latency_list)
    log_callback(f"keyboard average clicking latency:{average_latency:.3f} ms", False)

    return True if average_latency <= threshold else False


def serial_test(ser: serial.Serial):
    for _ in range(10):
        start = time.perf_counter()  # mark the time
        ser.write(CMD_test + b"\n")  # example: send 'C' = click command
        ser.flush()
        ser.read()
        end = time.perf_counter()
        print(f"serial port trans time {(end - start)/2:.6f} seconds")


def mouse_move_to_safe_place():
    # move mouse to safe place (top-right) before clicking
    screen_width, screen_height = pyautogui.size()
    pyautogui.click(x=screen_width - 10, y=10)


def run_test(test_case: str, b_config: Basic_Config, log_callback) -> bool:
    """_summary_
    Args:
        test_case (str): test case
        b_config (Basic_Config): the configation of the test
        log_callback (_type_): present the log one textbox and save the log txt

    Returns:
        bool: _description_
    """
    # connect to arduino board
    global g_COM_PORT
    res = True
    g_COM_PORT = b_config.com
    ser: serial.Serial = None

    # do nothing if test_case = Idle
    if test_case == Test_case.Idle.value:
        log_callback(f"waiting for {b_config.wake_up_time_s} sec...", False)
        time.sleep(b_config.wake_up_time_s)
        return True

    if not g_COM_PORT:
        try:
            g_COM_PORT = get_arduino_port(
                port=b_config.target_port_desc, log_callback=log_callback
            )
        except:
            log_callback("Can not find the arduino board!", False)
            return False

    try:
        if wait_for_port(g_COM_PORT, 30):
            ser = serial.Serial(g_COM_PORT, 115200, timeout=30)
            time.sleep(5)
        else:
            log_callback(f"Cant not find the serial device", False)
            return False

    except Exception as ex:
        log_callback(f"Serial port connection error:{ex}", False)
        return False

    log_callback(f"test case: {test_case}", False)

    match test_case:

        # case 'Idle' :
        # log_callback(f"waiting for {b_config.wake_up_time_s} sec...")
        # dut_states_init(power_states=Power_States.idle, wake_up_time_s=b_config.wake_up_time_s,
        # sleep_time_s=b_config.sleep_time_s,ser=ser)

        case Test_case.MS.value:
            # move to safe place
            mouse_move_to_safe_place()
            log_callback(f"go to MS states for {b_config.sleep_time_s} sec...", False)
            dut_states_init(
                power_states=Power_States.go_to_s3,
                wake_up_time_s=b_config.wake_up_time_s,
                sleep_time_s=b_config.sleep_time_s,
                ser=ser,
            )
            # make sure the laptop go to MS states
            time.sleep(10)

        case Test_case.S4.value:
            log_callback(f"go to s4 states for {b_config.sleep_time_s} sec...", False)
            dut_states_init(
                power_states=Power_States.go_to_s4,
                wake_up_time_s=b_config.wake_up_time_s,
                sleep_time_s=b_config.sleep_time_s,
                ser=ser,
            )

        case Test_case.Mouse_function.value:
            # mouse move to save place
            mouse_move_to_safe_place()
            # mouse function test
            res = mouse_keyboard_function_detect(
                ser=ser,
                command=CMD_mouse_clicking,
                timeout_s=b_config.timeout_s,
                log_callback=log_callback,
            )
            if not res:
                log_callback("mouse function test fail!", False)
                log_callback("dump wrt log...", False)
                WRTController.dump_wrt_log(log_path=b_config.report_path)

            time.sleep(5)

        case Test_case.keyboard_function.value:
            # mouse function test
            res = mouse_keyboard_function_detect(
                ser=ser,
                command=CMD_keyboard_clicking,
                timeout_s=b_config.timeout_s,
                log_callback=log_callback,
            )
            if not res:
                log_callback("keyboard function test fail!", False)
                log_callback("dump wrt log...", False)
                WRTController.dump_wrt_log(log_path=b_config.report_path)
            time.sleep(5)

        case Test_case.Mouse_random.value:
            # move to safe place
            mouse_move_to_safe_place()
            # mouse random click
            res = mouse_keyboard_random_click(
                ser=ser,
                command=CMD_mouse_random_clicking,
                timeout_s=b_config.timeout_s,
                log_callback=log_callback,
            )
            if not res:
                log_callback("mouse random test fail!", False)
                log_callback("dump wrt log...", False)
                WRTController.dump_wrt_log(log_path=b_config.report_path)
            time.sleep(5)

        case Test_case.Keyboard_random.value:
            # mouse function test
            res = mouse_keyboard_random_click(
                ser=ser,
                command=CMD_keyboard_random_clicking,
                timeout_s=b_config.timeout_s,
                log_callback=log_callback,
            )
            if not res:
                log_callback("keyboard random test fail!", False)
                log_callback("dump wrt log...", False)
                WRTController.dump_wrt_log(log_path=b_config.report_path)
            time.sleep(5)

        case Test_case.Headset_init.value:
            # headset init
            res = headset_init(
                headset=b_config.headset,
                headset_states=b_config.headset_setting,
                timeout_s=b_config.timeout_s,
                ser=ser,
                command=CMD_servo,
            )
            if not res:
                log_callback(
                    "Can not turn on the Headset or headset can not connect to the dut, stop testing!",
                    False,
                )
                log_callback("dump wrt log...", False)
                WRTController.dump_wrt_log(log_path=b_config.report_path)
            else:
                log_callback("Turn on the headset successfully, connected", False)
            time.sleep(5)

        case Test_case.Headset_input.value:

            res = headset_input_test(
                b_config=b_config, ser=ser, log_callback=log_callback
            )

        case Test_case.Headset_output.value:
            res = headset_output_test(
                b_config=b_config, ser=ser, log_callback=log_callback
            )

        case Test_case.Headset_del.value:
            res = headset_del(
                headset=b_config.headset,
                headset_states=b_config.headset_setting,
                timeout_s=b_config.timeout_s,
                ser=ser,
                command=CMD_servo,
            )
            if not res:
                log_callback("Can not turn off the Headset, stop test!", False)
            else:
                log_callback("Headset turn off successfully, disconneted", False)

        case Test_case.Mouse_latency.value:
            # move to safe place
            mouse_move_to_safe_place()

            res = mouse_latency(
                ser=ser,
                threshold=80,
                timeout_s=b_config.timeout_s,
                log_callback=log_callback,
            )

        case Test_case.keyboard_latency.value:
            res = keyboard_latency(
                ser=ser,
                threshold=100,
                timeout_s=b_config.timeout_s,
                log_callback=log_callback,
            )

        case Test_case.Environment_init.value:
            t_control = MeetingControl(
                meeting_link=b_config.teams_url, teams_path="\\teams_call\\"
            )
            v_control = VPTControl()
            res = env_init(
                ENV_source=b_config.ENV_source,
                t_control=t_control,
                v_control=v_control,
                log_callback=log_callback,
            )

        case Test_case.Environment_restore.value:
            res = env_restore(env_source=b_config.ENV_source, log_callback=log_callback)

        case "test":
            serial_test(ser=ser)

        case _:
            log_callback("Not match any test case, please check!", False)
            res = False

    # close serial connect after testing:
    try:
        print("Relsease serial port!")
        ser.close()
    except:
        pass

    return res

    # summary the test result
    """
    if b_config.do_mouse_flag: log_callback(f'mouse function: {res_mouse}')
    if b_config.do_headset_output_flag: log_callback(f'headset output function: {res_output}, init issue:{output_init_flag}')
    if b_config.do_headset_input_flag: log_callback(f'headset input function: {res_input}')
    if (res_output or not b_config.do_headset_output_flag) and (res_input or not b_config.do_headset_input_flag) and (res_mouse or not b_config.do_mouse_flag) :
        test_success_count+=1
        continue_fail_count = 0
    else:
        continue_fail_count+=1
    test_total_count+=1
    log_callback(f'successfully test  {test_success_count} times')
    log_callback(f'Total test  {test_total_count} times')
    if continue_fail_count >= b_config.continue_fail_limit:
        log_callback("out of maxium continue fail range, stop testing!")
        log_callback(f"first issue occuring time: {issue_occuring_time}")
        break
     """


if __name__ == "__main__":
    """
    b_config  = Basic_Config()
    def log_callback(meg:str):
        print(meg)
        logger.info(meg)
    test_case = 'keyboard_function'
    run_test(test_case,b_config,log_callback=log_callback)
    """
    print(gw.getAllTitles())
